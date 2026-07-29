"""Read-only Excel staging and reconciliation for controlled legacy migration."""

from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Mapping


def json_value(value):
    """Keep source values (including formulas) serializable and traceable."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalized_header(value, column_index: int) -> str:
    text = " ".join(str(value or "").strip().split())
    return text or f"column_{column_index}"


def find_header_row(rows) -> tuple[int, list] | tuple[None, list]:
    """Find the first likely heading row without trusting fixed template offsets."""
    for row_number, row in enumerate(rows, 1):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
        text_cells = [cell for cell in non_empty if isinstance(cell, str) and not cell.startswith("=")]
        if len(non_empty) >= 2 and len(text_cells) >= 2:
            return row_number, list(row)
    return None, []


def build_row_payload(filename: str, sheet_name: str, row_number: int, headers, values) -> dict:
    fields = {}
    cells = {}
    for column, (header, value) in enumerate(zip(headers, values), 1):
        if value is None:
            continue
        key = normalized_header(header, column)
        while key in fields:
            key = f"{key}_{column}"
        fields[key] = json_value(value)
        cells[column] = json_value(value)
    return {
        "source": {
            "filename": filename,
            "sheet": sheet_name,
            "row": row_number,
            "cells": cells,
        },
        "values": fields,
    }


def reconciliation_from_payloads(payloads: list[dict]) -> dict:
    """Return transparent row counts and source numeric totals; never calculate formulas."""
    totals = defaultdict(float)
    for payload in payloads:
        for header, value in payload["values"].items():
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                totals[header] += float(value)
    return {"row_count": len(payloads), "numeric_totals": dict(sorted(totals.items()))}


class ExcelImportService:
    """Stages workbooks first; importing into operational tables is explicitly approved."""

    def __init__(self, db_instance):
        self.db = db_instance

    def stage_workbook(self, filename, actor_username="system") -> dict:
        """Store immutable source rows with filename/sheet/row provenance.

        The workbook is opened read-only.  Formula cells are retained as formulas,
        and no accounting table is altered during this staging step.
        """
        path = Path(filename)
        if not path.is_file():
            raise FileNotFoundError(path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Only .xlsx and .xlsm archives are accepted for controlled staging.")
        digest = source_sha256(path)
        existing = self.db.fetch_one(
            "SELECT * FROM Import_Batches WHERE source_sha256 = %s", (digest,)
        )
        if existing:
            return {"reused": True, "batch_id": existing["id_batch"], "status": existing["status"]}

        try:
            from openpyxl import load_workbook
        except ImportError as error:  # Clear deployment failure rather than silent partial import.
            raise RuntimeError("openpyxl is required to stage legacy Excel files.") from error

        workbook = load_workbook(path, read_only=True, data_only=False)
        success, batch_id = self.db.execute(
            """INSERT INTO Import_Batches (source_filename, source_sha256, imported_by)
               VALUES (%s, %s, %s)""",
            (path.name, digest, actor_username),
        )
        if not success:
            raise RuntimeError("Unable to create the import batch.")

        reconciliation = {"filename": path.name, "sheets": [], "rejected_rows": []}
        for sheet in workbook.worksheets:
            sampled_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True))
            header_row, headers = find_header_row(sampled_rows)
            if header_row is None:
                reconciliation["rejected_rows"].append(
                    {"sheet": sheet.title, "reason": "No reliable header row was found."}
                )
                continue
            payloads = []
            for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                payload = build_row_payload(path.name, sheet.title, row_number, headers, row)
                if not payload["values"]:
                    continue
                payloads.append(payload)
                self.db.execute(
                    """INSERT INTO Import_Rows
                       (batch_id, sheet_name, source_row, source_reference, entity_type, payload_json)
                       VALUES (%s, %s, %s, %s, %s, %s)""",
                    (
                        batch_id, sheet.title, row_number, f"{sheet.title}!{row_number}",
                        "LEGACY_EXCEL_ROW", json.dumps(payload, ensure_ascii=False),
                    ),
                )
            sheet_summary = reconciliation_from_payloads(payloads)
            sheet_summary.update({"sheet": sheet.title, "header_row": header_row})
            reconciliation["sheets"].append(sheet_summary)

        self.db.execute(
            "UPDATE Import_Batches SET reconciliation_json = %s WHERE id_batch = %s",
            (json.dumps(reconciliation, ensure_ascii=False), batch_id),
        )
        return {"reused": False, "batch_id": batch_id, "status": "STAGED", "reconciliation": reconciliation}

    def validate_batch(self, batch_id, actor_username, notes=None):
        """Explicit human approval is required before operational-row handlers run."""
        batch = self.db.fetch_one("SELECT * FROM Import_Batches WHERE id_batch = %s", (batch_id,))
        if not batch or batch["status"] != "STAGED":
            raise ValueError("Only staged import batches can be validated.")
        success, _ = self.db.execute(
            "UPDATE Import_Batches SET status = 'VALIDATED' WHERE id_batch = %s AND status = 'STAGED'",
            (batch_id,),
        )
        if success:
            self.db.execute(
                """INSERT INTO Audit_Events (actor_username, action_code, entity_type, entity_id, reason)
                   VALUES (%s, 'IMPORT_BATCH_VALIDATED', 'Import_Batches', %s, %s)""",
                (actor_username, str(batch_id), notes),
            )
        return success

    def import_validated_batch(
        self, batch_id, actor_username, row_handler: Callable[[Mapping], tuple[bool, str | None, str | None]],
    ) -> dict:
        """Apply reviewed rows through a domain handler, preserving rejects and retries.

        ``row_handler`` must call the financial/HR service appropriate to the row;
        direct generic table writes are intentionally not provided here.
        """
        batch = self.db.fetch_one("SELECT * FROM Import_Batches WHERE id_batch = %s", (batch_id,))
        if not batch or batch["status"] not in {"VALIDATED", "IMPORTED"}:
            raise ValueError("The batch must be validated before import.")
        rows = self.db.fetch_all(
            "SELECT * FROM Import_Rows WHERE batch_id = %s AND status = 'STAGED' ORDER BY id_import_row",
            (batch_id,),
        )
        imported = rejected = 0
        for row in rows:
            payload = json.loads(row["payload_json"])
            accepted, entity_id, rejection_reason = row_handler(payload)
            if accepted:
                self.db.execute(
                    "UPDATE Import_Rows SET status = 'IMPORTED', entity_id = %s WHERE id_import_row = %s",
                    (str(entity_id) if entity_id is not None else None, row["id_import_row"]),
                )
                imported += 1
            else:
                self.db.execute(
                    "UPDATE Import_Rows SET status = 'REJECTED', rejection_reason = %s WHERE id_import_row = %s",
                    (rejection_reason or "Rejected by reviewed import handler.", row["id_import_row"]),
                )
                rejected += 1
        status = "IMPORTED" if rejected == 0 else "REJECTED"
        self.db.execute("UPDATE Import_Batches SET status = %s WHERE id_batch = %s", (status, batch_id))
        self.db.execute(
            """INSERT INTO Audit_Events (actor_username, action_code, entity_type, entity_id, new_values)
               VALUES (%s, 'IMPORT_BATCH_APPLIED', 'Import_Batches', %s, %s)""",
            (actor_username, str(batch_id), json.dumps({"imported": imported, "rejected": rejected})),
        )
        return {"batch_id": batch_id, "status": status, "imported": imported, "rejected": rejected}
